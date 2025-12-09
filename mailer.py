"""
Outlook 이메일 + Excel 보고서
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from config import EMAIL_CONFIG


def create_html(df, period_label):
    """
    이메일용 HTML 생성 - 통합 테이블
    """
    
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Malgun Gothic', Arial, sans-serif; font-size: 10pt; }}
            h2 {{ color: #00A651; border-bottom: 2px solid #00A651; padding-bottom: 10px; font-size: 14pt; }}
            table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; font-size: 10pt; }}
            th {{ background: #00A651; color: white; padding: 8px; text-align: left; font-size: 10pt; }}
            td {{ border: 1px solid #ddd; padding: 8px; vertical-align: top; font-size: 10pt; }}
            tr:nth-child(even) {{ background: #f9f9f9; }}
            a {{ color: #00A651; text-decoration: underline; cursor: pointer; }}
            .footer {{ margin-top: 30px; color: #888; font-size: 9pt; border-top: 1px solid #ddd; padding-top: 15px; }}
        </style>
    </head>
    <body>
        <h2>🧪 Polymer 뉴스 리포트 - {period_label}</h2>
        <p>PE · PP · PO · S-OIL · Aramco · Sabic 관련 주요 뉴스 (Top {len(df)})</p>
    """
    
    if not df.empty:
        html += """
        <table>
            <tr>
                <th style="width:3%">No.</th>
                <th style="width:10%">Main</th>
                <th style="width:8%">Bonus</th>
                <th style="width:25%">제목</th>
                <th style="width:49%">요약</th>
                <th style="width:5%">링크</th>
            </tr>
        """
        
        for idx, row in df.iterrows():
            title = row.get("title", "")
            link = row.get("link", "#")
            main_kw = row.get("main_keywords", "")
            bonus_kw = row.get("bonus_keywords", "")
            summary = row.get("summary", "")
            
            html += f"""
            <tr>
                <td style="text-align:center">{idx + 1}</td>
                <td>{main_kw}</td>
                <td>{bonus_kw}</td>
                <td>{title}</td>
                <td>{summary}</td>
                <td style="text-align:center"><a href="{link}">보기</a></td>
            </tr>
            """
        
        html += "</table>"
    
    html += f"""
        <div class="footer">
            <p>🏢 Polymer 영업부문 | Aramco Subsidiary | Sabic Partner</p>
            <p>생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
        </div>
    </body>
    </html>
    """
    
    return html


def send_outlook(df, period_label, to_email, cc_email=None):
    """
    Outlook으로 이메일 발송
    """
    try:
        import win32com.client as win32
        import pythoncom
        pythoncom.CoInitialize()
        
        outlook = win32.Dispatch("outlook.application")
        mail = outlook.CreateItem(0)
        
        mail.To = to_email
        if cc_email:
            mail.CC = cc_email
        
        mail.Subject = f"{EMAIL_CONFIG['subject_prefix']} {period_label} 주요 동향"
        mail.HTMLBody = create_html(df, period_label)
        
        mail.Send()
        print(f"✅ 이메일 발송 완료: {to_email}")
        return True
        
    except Exception as e:
        print(f"❌ 이메일 발송 실패: {e}")
        return False


def save_draft(df, period_label, to_email=""):
    """
    Outlook 임시보관함에 저장
    """
    try:
        import win32com.client as win32
        import pythoncom
        pythoncom.CoInitialize()
        
        outlook = win32.Dispatch("outlook.application")
        mail = outlook.CreateItem(0)
        
        mail.To = to_email
        mail.Subject = f"{EMAIL_CONFIG['subject_prefix']} {period_label} 주요 동향"
        mail.HTMLBody = create_html(df, period_label)
        
        mail.Save()
        print("✅ 임시보관함 저장 완료")
        return True
        
    except Exception as e:
        print(f"❌ 저장 실패: {e}")
        return False


def save_excel_report(df, period_label, save_path="./result"):
    """
    Excel 보고서 저장
    """
    os.makedirs(save_path, exist_ok=True)
    
    wb = Workbook()
    
    # 시트1: 뉴스 리포트
    ws1 = wb.active
    ws1.title = "NewsReport"
    
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    columns = ["category", "keyword", "title", "summary", "source", "date", "strategy_score", "link"]
    available_cols = [c for c in columns if c in df.columns]
    export_df = df[available_cols].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
    
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 45
    ws1.column_dimensions['D'].width = 55
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 40
    
    # 시트2: 카테고리 통계
    ws2 = wb.create_sheet("Summary")
    ws2.append(["카테고리", "기사 수", "평균 점수", "주요 키워드"])
    
    for category in df["category"].unique():
        cat_df = df[df["category"] == category]
        avg_score = cat_df["strategy_score"].mean() if "strategy_score" in cat_df.columns else 0
        top_kw = cat_df["keyword"].value_counts().index[0] if len(cat_df) > 0 else ""
        ws2.append([category, len(cat_df), round(avg_score, 2), top_kw])
    
    filename = f"{period_label.replace('/', '-').replace(' ', '_')}_polymer_report.xlsx"
    filepath = os.path.join(save_path, filename)
    wb.save(filepath)
    
    print(f"📊 Excel 저장: {filepath}")
    return filepath